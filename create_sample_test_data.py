"""
Script to create sample test data Excel files for the API Framework
Run this script to generate film_test_data.xlsx and film_xml_test_data.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json
import os

def create_json_test_data():
    """Create sample JSON test data Excel file"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Films"

    # Define headers
    headers = [
        'title',
        'description',
        'release_year',
        'language_id',
        'rental_duration',
        'rental_rate',
        'length',
        'replacement_cost',
        'rating',
        'special_features',
        'fulltext',
        'expected_response'
    ]

    # Add headers with styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample test data rows
    test_data_rows = [
        [
            'The Shawshank Redemption',
            'A drama film about hope and survival',
            2023,
            1,  # English
            4,  # rental_duration
            4.99,
            142,
            29.99,
            'PG-13',
            'Behind the Scenes',
            'drama redemption prison',
            '{"film_id": 1, "title": "The Shawshank Redemption", "rental_rate": 4.99}'
        ],
        [
            'The Dark Knight',
            'A superhero action film',
            2023,
            1,  # English
            5,
            5.99,
            152,
            34.99,
            'PG-13',
            'Behind the Scenes,Trailers',
            'batman action superhero',
            '{"film_id": 2, "title": "The Dark Knight", "rental_rate": 5.99}'
        ],
        [
            'Inception',
            'A science fiction thriller',
            2023,
            1,  # English
            4,
            6.99,
            148,
            39.99,
            'PG-13',
            'Behind the Scenes,Commentary',
            'scifi dreams thriller',
            '{"film_id": 3, "title": "Inception", "rental_rate": 6.99}'
        ],
        [
            'Forrest Gump',
            'A heartwarming drama',
            2023,
            1,  # English
            3,
            3.99,
            142,
            24.99,
            'PG',
            'Behind the Scenes',
            'drama life story',
            '{"film_id": 4, "title": "Forrest Gump", "rental_rate": 3.99}'
        ]
    ]

    # Add data rows
    for row_num, row_data in enumerate(test_data_rows, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Adjust column widths
    widths = [20, 30, 12, 12, 16, 13, 8, 18, 10, 25, 25, 50]
    for col_num, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = width

    # Save file
    file_path = os.path.join(os.path.dirname(__file__), 'test_data', 'film_test_data.xlsx')
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    workbook.save(file_path)
    print(f"✓ Created: {file_path}")


def create_xml_test_data():
    """Create sample XML test data Excel file"""
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Films"

    # Define headers
    headers = [
        'title',
        'release_year',
        'language_id',
        'rental_duration',
        'rental_rate',
        'length',
        'replacement_cost',
        'rating',
        'expected_response'
    ]

    # Add headers with styling
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample test data rows
    test_data_rows = [
        [
            'The Matrix',
            2023,
            1,  # English
            4,
            5.99,
            136,
            34.99,
            'R',
            '{"film_id": 10, "title": "The Matrix", "rental_rate": 5.99}'
        ],
        [
            'Titanic',
            2023,
            1,  # English
            5,
            4.99,
            194,
            29.99,
            'PG-13',
            '{"film_id": 11, "title": "Titanic", "rental_rate": 4.99}'
        ]
    ]

    # Add data rows
    for row_num, row_data in enumerate(test_data_rows, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Adjust column widths
    widths = [20, 12, 12, 16, 13, 8, 18, 10, 50]
    for col_num, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(col_num)].width = width

    # Save file
    file_path = os.path.join(os.path.dirname(__file__), 'test_data', 'film_xml_test_data.xlsx')
    workbook.save(file_path)
    print(f"✓ Created: {file_path}")


if __name__ == "__main__":
    print("Creating sample test data Excel files...")
    create_json_test_data()
    create_xml_test_data()
    print("\nSample data files created successfully!")
    print("You can now modify these Excel files with your actual test data.")

